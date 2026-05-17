"""
Greenpack Pro v2.0 — Multi-Up Report Generator
PDF + Excel QC reports for multi-label sheets.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.config import get_settings

log = logging.getLogger(__name__)
settings = get_settings()


def generate_multi_up_pdf(job_id: str, config: dict, result: dict) -> Optional[Path]:
    """Generate PDF report for a multi-up inspection"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            Image as RLImage, PageBreak, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_multi_up_report.pdf"

        NAVY = colors.HexColor("#0D1B2A")
        CYAN = colors.HexColor("#00C2CB")
        GREEN = colors.HexColor("#22A06B")
        RED = colors.HexColor("#E5383B")
        LIGHT = colors.HexColor("#F0F6FF")
        SILVER = colors.HexColor("#E8EEF4")

        def ps(name, **kw):
            d = dict(fontName="Helvetica", fontSize=10, leading=14)
            d.update(kw)
            return ParagraphStyle(name, **d)

        h1 = ps("h1", fontName="Helvetica-Bold", fontSize=14, textColor=NAVY)
        body = ps("body", fontSize=9.5, textColor=colors.HexColor("#333333"), leading=13)

        sheet_pass = result.get("sheet_pass", False)
        overall_score = result.get("overall_score", 0)
        labels_found = result.get("labels_found", 0)
        labels_expected = result.get("labels_expected")
        labels_missing = result.get("labels_missing", 0)
        labels_passed = result.get("labels_passed", 0)
        labels_failed = result.get("labels_failed", 0)
        pass_rate = result.get("pass_rate", 0)

        def page_template(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(NAVY)
            canvas.rect(0, A4[1] - 30, A4[0], 30, fill=1, stroke=0)
            canvas.setFont("Helvetica-Bold", 9)
            canvas.setFillColor(colors.white)
            canvas.drawString(1.5*cm, A4[1] - 20, "Greenpack Pro — Multi-Up Sheet QC Report")
            canvas.drawRightString(A4[0] - 1.5*cm, A4[1] - 20,
                                    f"Job: {config.get('job_ref', job_id[:8])}")
            canvas.setFillColor(SILVER)
            canvas.rect(0, 0, A4[0], 20, fill=1, stroke=0)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#666666"))
            canvas.drawString(1.5*cm, 6,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Greenpack Pro v2.0 (Multi-Up)")
            canvas.drawRightString(A4[0] - 1.5*cm, 6, f"Page {doc.page}")
            canvas.restoreState()

        doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=1.5*cm)
        story = []

        # ── Cover ──────────────────────────────────────────────────────────────
        cover_title = Paragraph(
            f'<font color="#FFFFFF"><b>Multi-Up Sheet Inspection — {config.get("product_name", "Label")}</b></font>',
            ps("cover", fontName="Helvetica-Bold", fontSize=16, textColor=colors.white, alignment=TA_CENTER)
        )
        cover = Table([[cover_title]], colWidths=[A4[0] - 3*cm])
        cover.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), NAVY),
                                    ("TOPPADDING", (0,0), (-1,-1), 12),
                                    ("BOTTOMPADDING", (0,0), (-1,-1), 12)]))
        story.append(cover)
        story.append(Spacer(1, 15))

        # ── Summary ────────────────────────────────────────────────────────────
        status_color = GREEN if sheet_pass else RED
        status_text = "✓ SHEET PASS" if sheet_pass else "✗ SHEET FAIL"

        score_cell = Paragraph(
            f'{overall_score:.1f}',
            ps("score", fontName="Helvetica-Bold", fontSize=36, alignment=TA_CENTER,
               textColor=status_color)
        )
        status_cell = Paragraph(
            f'<b><font color="{"#22A06B" if sheet_pass else "#E5383B"}">{status_text}</font></b>',
            ps("stat", fontName="Helvetica-Bold", fontSize=20, alignment=TA_CENTER)
        )

        score_tbl = Table([[score_cell, status_cell]], colWidths=[6*cm, 12*cm])
        score_tbl.setStyle(TableStyle([
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("BOX", (0,0), (-1,-1), 1, SILVER),
            ("TOPPADDING", (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ]))
        story.append(score_tbl)
        story.append(Spacer(1, 12))

        # ── Key metrics ────────────────────────────────────────────────────────
        metrics = [
            [Paragraph("Labels Found", ps("k", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Expected", ps("k", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Missing", ps("k", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Passed", ps("k", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Failed", ps("k", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Pass Rate", ps("k", fontName="Helvetica-Bold", fontSize=9))],
            [Paragraph(str(labels_found), ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER)),
             Paragraph(str(labels_expected or "—"), ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER)),
             Paragraph(str(labels_missing), ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER,
                       textColor=RED if labels_missing else colors.black)),
             Paragraph(str(labels_passed), ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER,
                       textColor=GREEN)),
             Paragraph(str(labels_failed), ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER,
                       textColor=RED if labels_failed else colors.black)),
             Paragraph(f"{pass_rate:.1f}%", ps("v", fontSize=16, fontName="Helvetica-Bold", alignment=TA_CENTER,
                       textColor=GREEN if pass_rate >= 95 else RED))],
        ]
        mtbl = Table(metrics, colWidths=[(A4[0] - 3*cm) / 6] * 6)
        mtbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("GRID", (0,0), (-1,-1), 0.5, SILVER),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ]))
        story.append(mtbl)
        story.append(Spacer(1, 16))

        # ── Sheet details ──────────────────────────────────────────────────────
        story.append(Paragraph("Sheet Information", h1))
        story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
        story.append(Spacer(1, 8))
        details = [
            ["Job Reference:", config.get("job_ref", job_id[:12]),
             "Imposition:", result.get("imposition", "—")],
            ["Client:", config.get("client_name", "—"),
             "Detection Method:", result.get("detection_method", "—")],
            ["Product:", config.get("product_name", "—"),
             "Detection Confidence:", f"{result.get('detection_confidence', 0) * 100:.1f}%"],
            ["Inspector:", config.get("inspector_name", "—"),
             "Processing Time:", f"{result.get('processing_time_ms', 0) / 1000:.1f}s"],
            ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M"),
             "Transparent Labels:", "Yes" if config.get("is_transparent") else "No"],
        ]
        dtbl = Table(details, colWidths=[3*cm, 5.5*cm, 3.5*cm, 5*cm])
        dtbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, LIGHT]),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(dtbl)
        story.append(Spacer(1, 16))

        # ── Per-label results table ────────────────────────────────────────────
        per_label = result.get("per_label_results", [])
        if per_label:
            story.append(Paragraph(f"Per-Label Results ({len(per_label)} labels)", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))

            header = [["Label", "Score", "OCR", "Color", "SSIM", "Barcode", "Reg.", "Die-cut", "Mottling", "Status"]]
            rows = []
            for lb in per_label:
                status_icon = "✓ PASS" if lb.get("pass_fail") else "✗ FAIL"
                rows.append([
                    lb.get("label_id", "?"),
                    f'{lb.get("overall_score", 0):.0f}',
                    f'{lb.get("ocr_score", 0):.0f}',
                    f'{lb.get("color_score", 0):.0f}',
                    f'{lb.get("ssim_score", 0):.0f}',
                    f'{lb.get("barcode_score", 0):.0f}',
                    f'{lb.get("registration_score", 0):.0f}',
                    f'{lb.get("die_cut_score", 0):.0f}',
                    f'{lb.get("mottling_score", 0):.0f}',
                    status_icon,
                ])

            tbl = Table(header + rows, colWidths=[1.4*cm]*9 + [1.9*cm])
            style_cmds = [
                ("BACKGROUND", (0,0), (-1,0), NAVY),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 8),
                ("GRID", (0,0), (-1,-1), 0.3, SILVER),
                ("ALIGN", (1,1), (-1,-1), "CENTER"),
                ("TOPPADDING", (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ]
            for i, lb in enumerate(per_label, 1):
                bg = colors.HexColor("#FEE2E2") if not lb.get("pass_fail") else (colors.white if i % 2 else LIGHT)
                style_cmds.append(("BACKGROUND", (0,i), (-1,i), bg))

            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)
            story.append(Spacer(1, 12))

        # ── Annotated sheet image ──────────────────────────────────────────────
        sheet_img_path = result.get("sheet_annotated_path")
        if sheet_img_path and Path(sheet_img_path).exists():
            story.append(PageBreak())
            story.append(Paragraph("Annotated Sheet — Per-Label Pass/Fail", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))
            story.append(RLImage(str(sheet_img_path),
                                  width=A4[0] - 3*cm, height=(A4[0] - 3*cm) * 0.5))

        # ── Failed labels detail section ───────────────────────────────────────
        failed_labels = [lb for lb in per_label if not lb.get("pass_fail")]
        if failed_labels:
            story.append(PageBreak())
            story.append(Paragraph(f"Failed Labels Detail ({len(failed_labels)} labels)", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))

            for lb in failed_labels[:20]:
                label_id = lb.get("label_id", "?")
                story.append(Paragraph(
                    f"<b>Label {label_id} — Score: {lb.get('overall_score', 0):.1f}</b>",
                    ps("ld", fontName="Helvetica-Bold", fontSize=10, textColor=RED)
                ))

                # OCR errors
                ocr_errs = lb.get("ocr_errors", [])
                if ocr_errs:
                    err_text = f"<b>OCR Errors ({len(ocr_errs)}):</b> "
                    err_text += "; ".join([f"'{e.get('master_text', '')[:15]}' → '{e.get('scan_text', '')[:15]}'"
                                            for e in ocr_errs[:3]])
                    story.append(Paragraph(err_text, body))

                # Defects
                defects = lb.get("defects", [])
                if defects:
                    story.append(Paragraph(
                        f"<b>Defects:</b> {len(defects)} — {', '.join(d.get('type', '?') for d in defects[:5])}",
                        body
                    ))

                # Registration
                reg = lb.get("registration", {})
                if reg and reg.get("offset_px", 0) > 5:
                    story.append(Paragraph(
                        f"<b>Registration drift:</b> {reg.get('offset_px', 0):.1f}px offset",
                        body
                    ))

                story.append(Spacer(1, 6))

        doc.build(story, onFirstPage=page_template, onLaterPages=page_template)
        log.info(f"Multi-up PDF report generated: {out_path}")
        return out_path

    except Exception as e:
        log.exception(f"Multi-up PDF report generation failed: {e}")
        return None


def generate_multi_up_excel(job_id: str, config: dict, result: dict) -> Optional[Path]:
    """Generate multi-sheet Excel report for a multi-up inspection"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_multi_up_results.xlsx"

        wb = openpyxl.Workbook()
        header_fill = PatternFill("solid", fgColor="0D1B2A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        pass_fill = PatternFill("solid", fgColor="D1FAE5")
        fail_fill = PatternFill("solid", fgColor="FEE2E2")

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        summary = [
            ["Job Reference", config.get("job_ref", job_id[:12])],
            ["Client", config.get("client_name", "")],
            ["Product", config.get("product_name", "")],
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Mode", "Multi-Up Sheet"],
            ["", ""],
            ["Labels Found", result.get("labels_found", 0)],
            ["Labels Expected", result.get("labels_expected") or "Not specified"],
            ["Labels Missing", result.get("labels_missing", 0)],
            ["Labels Passed", result.get("labels_passed", 0)],
            ["Labels Failed", result.get("labels_failed", 0)],
            ["Pass Rate", f"{result.get('pass_rate', 0):.1f}%"],
            ["", ""],
            ["Overall Score", f"{result.get('overall_score', 0):.1f}"],
            ["Sheet Pass/Fail", "PASS" if result.get("sheet_pass") else "FAIL"],
            ["Imposition", result.get("imposition", "—")],
            ["Detection Method", result.get("detection_method", "—")],
            ["Processing Time (ms)", result.get("processing_time_ms", 0)],
        ]
        for row_idx, (k, v) in enumerate(summary, 1):
            ws.cell(row_idx, 1, k).font = Font(bold=True)
            ws.cell(row_idx, 2, str(v))
            if k == "Sheet Pass/Fail":
                fill = pass_fill if v == "PASS" else fail_fill
                ws.cell(row_idx, 2).fill = fill

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        # ── Per-label sheet ────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Per-Label Results")
        headers = ["Label ID", "Row", "Column", "Overall Score", "Pass/Fail",
                   "OCR Score", "Color Score", "SSIM Score", "Barcode Score",
                   "Registration", "Die-Cut", "Mottling",
                   "Alignment Conf.", "OCR Errors", "Defects", "Position X", "Position Y"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        per_label = result.get("per_label_results", [])
        for i, lb in enumerate(per_label, 2):
            bbox = lb.get("bbox", {})
            ws2.cell(i, 1, lb.get("label_id"))
            ws2.cell(i, 2, lb.get("row", 0) + 1)
            ws2.cell(i, 3, lb.get("col", 0) + 1)
            ws2.cell(i, 4, lb.get("overall_score"))
            ws2.cell(i, 5, "PASS" if lb.get("pass_fail") else "FAIL")
            ws2.cell(i, 6, lb.get("ocr_score"))
            ws2.cell(i, 7, lb.get("color_score"))
            ws2.cell(i, 8, lb.get("ssim_score"))
            ws2.cell(i, 9, lb.get("barcode_score"))
            ws2.cell(i, 10, lb.get("registration_score"))
            ws2.cell(i, 11, lb.get("die_cut_score"))
            ws2.cell(i, 12, lb.get("mottling_score"))
            ws2.cell(i, 13, lb.get("alignment_confidence"))
            ws2.cell(i, 14, len(lb.get("ocr_errors", [])))
            ws2.cell(i, 15, len(lb.get("defects", [])))
            ws2.cell(i, 16, bbox.get("x", ""))
            ws2.cell(i, 17, bbox.get("y", ""))

            fill = pass_fill if lb.get("pass_fail") else fail_fill
            for c in range(1, 18):
                ws2.cell(i, c).fill = fill

        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]:
            ws2.column_dimensions[col_letter].width = 14

        # ── All defects sheet ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("All Defects")
        def_headers = ["Label", "Type", "Severity", "X", "Y", "Width", "Height", "Area (px²)"]
        for col, h in enumerate(def_headers, 1):
            cell = ws3.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill

        all_defects = result.get("all_defects", [])
        for i, d in enumerate(all_defects, 2):
            bbox = d.get("bbox", {})
            ws3.cell(i, 1, d.get("label_id", "?"))
            ws3.cell(i, 2, d.get("type", ""))
            ws3.cell(i, 3, d.get("severity", ""))
            ws3.cell(i, 4, bbox.get("x", ""))
            ws3.cell(i, 5, bbox.get("y", ""))
            ws3.cell(i, 6, bbox.get("w", ""))
            ws3.cell(i, 7, bbox.get("h", ""))
            ws3.cell(i, 8, d.get("area_pixels", ""))

        wb.save(str(out_path))
        log.info(f"Multi-up Excel report generated: {out_path}")
        return out_path

    except Exception as e:
        log.exception(f"Multi-up Excel report generation failed: {e}")
        return None
