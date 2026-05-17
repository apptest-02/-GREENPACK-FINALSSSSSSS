"""
Greenpack Pro v3.0 — Prepress Report Generator

PDF + Excel reports for prepress trial-vs-final comparisons.
Includes: GO/NO-GO decision, accuracy scores per trial, error breakdown,
waste savings estimate, annotated images.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.config import get_settings

log = logging.getLogger(__name__)
settings = get_settings()


def generate_prepress_pdf(job_id: str, config: dict, result: dict) -> Optional[Path]:
    """Generate prepress PDF report"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            Image as RLImage, PageBreak, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_prepress_report.pdf"

        NAVY = colors.HexColor("#0D1B2A")
        PURPLE = colors.HexColor("#7C3AED")
        GREEN = colors.HexColor("#22A06B")
        RED = colors.HexColor("#E5383B")
        ORANGE = colors.HexColor("#F59E0B")
        LIGHT = colors.HexColor("#F0F6FF")
        SILVER = colors.HexColor("#E8EEF4")

        def ps(name, **kw):
            d = dict(fontName="Helvetica", fontSize=10, leading=14)
            d.update(kw)
            return ParagraphStyle(name, **d)

        h1 = ps("h1", fontName="Helvetica-Bold", fontSize=14, textColor=NAVY)
        body = ps("body", fontSize=9.5, leading=13)

        decision = result.get("decision", "UNKNOWN")
        if decision == "GO":
            decision_color = GREEN
            decision_text = "✓ APPROVED FOR PRODUCTION"
        elif decision == "HOLD":
            decision_color = ORANGE
            decision_text = "⚠ HOLD — REVIEW REQUIRED"
        else:
            decision_color = RED
            decision_text = "✗ DO NOT PRINT — FIX ERRORS"

        def page_template(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(PURPLE)
            canvas.rect(0, A4[1] - 30, A4[0], 30, fill=1, stroke=0)
            canvas.setFont("Helvetica-Bold", 9)
            canvas.setFillColor(colors.white)
            canvas.drawString(1.5*cm, A4[1] - 20, "Greenpack Pro v3.0 — Prepress Comparison Report")
            canvas.drawRightString(A4[0] - 1.5*cm, A4[1] - 20,
                                    f"Job: {config.get('job_ref', job_id[:8])}")
            canvas.setFillColor(SILVER)
            canvas.rect(0, 0, A4[0], 20, fill=1, stroke=0)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#666666"))
            canvas.drawString(1.5*cm, 6,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Prepress / Real-Time Accuracy Report")
            canvas.drawRightString(A4[0] - 1.5*cm, 6, f"Page {doc.page}")
            canvas.restoreState()

        doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=1.5*cm)
        story = []

        # ── Cover ────────────────────────────────────────────────────────────
        cover_title = Paragraph(
            f'<font color="#FFFFFF"><b>PREPRESS COMPARISON — {config.get("product_name", "Print Job")}</b></font>',
            ps("cover", fontName="Helvetica-Bold", fontSize=15,
               textColor=colors.white, alignment=TA_CENTER)
        )
        cover = Table([[cover_title]], colWidths=[A4[0] - 3*cm])
        cover.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), PURPLE),
                                    ("TOPPADDING", (0,0), (-1,-1), 14),
                                    ("BOTTOMPADDING", (0,0), (-1,-1), 14)]))
        story.append(cover)
        story.append(Spacer(1, 16))

        # ── Decision banner ──────────────────────────────────────────────────
        decision_para = Paragraph(
            f'<font color="#FFFFFF"><b>{decision_text}</b></font>',
            ps("dec", fontName="Helvetica-Bold", fontSize=22, alignment=TA_CENTER,
               textColor=colors.white)
        )
        accuracy_para = Paragraph(
            f'<font color="#FFFFFF">Accuracy Score: <b>{result.get("accuracy_score", 0):.1f}%</b></font>',
            ps("acc", fontSize=14, alignment=TA_CENTER, textColor=colors.white)
        )
        reason_para = Paragraph(
            f'<font color="#FFFFFF"><i>{result.get("decision_reason", "")}</i></font>',
            ps("reason", fontSize=10, alignment=TA_CENTER, textColor=colors.white)
        )

        decision_tbl = Table([[decision_para], [accuracy_para], [reason_para]],
                              colWidths=[A4[0] - 3*cm])
        decision_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), decision_color),
            ("TOPPADDING", (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(decision_tbl)
        story.append(Spacer(1, 14))

        # ── Job details ──────────────────────────────────────────────────────
        story.append(Paragraph("Job Details", h1))
        story.append(HRFlowable(width="100%", thickness=1, color=PURPLE))
        story.append(Spacer(1, 6))
        details = [
            ["Job Reference:", config.get("job_ref", job_id[:12]),
             "Trials Inspected:", str(result.get("trial_count", 0))],
            ["Client:", config.get("client_name", "—"),
             "Best Trial:", f'{result.get("best_trial_score", 0):.1f}%'],
            ["Product:", config.get("product_name", "—"),
             "Worst Trial:", f'{result.get("worst_trial_score", 0):.1f}%'],
            ["Inspector:", config.get("inspector_name", "—"),
             "Processing Time:", f'{result.get("processing_time_ms", 0)/1000:.1f}s'],
            ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M"),
             "Min Required:", f'{config.get("min_accuracy_for_go", 90)}%'],
        ]
        dtbl = Table(details, colWidths=[3*cm, 5.5*cm, 3.5*cm, 5*cm])
        dtbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
            ("FONTNAME", (2,0), (2,-1), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, LIGHT]),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(dtbl)
        story.append(Spacer(1, 14))

        # ── Waste savings ────────────────────────────────────────────────────
        waste = result.get("waste_savings", {})
        if waste.get("estimated_savings_usd", 0) > 0:
            story.append(Paragraph("💰 Waste Prevention Estimate", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=PURPLE))
            story.append(Spacer(1, 6))

            savings_text = f'''<b>Estimated savings: <font color="#22A06B">${waste.get("estimated_savings_usd", 0):,.0f} USD</font></b><br/>
            <font size="9">By catching errors in trial proofs before full production, this inspection prevented an estimated
            <b>{waste.get("estimated_wasted_m2", 0):,.0f} m²</b> of wasted material out of a planned
            {waste.get("expected_run_m2", 0):,.0f} m² run.</font>'''
            story.append(Paragraph(savings_text, ps("waste", fontSize=11, leading=15)))
            story.append(Spacer(1, 14))

        # ── Per-trial scores ──────────────────────────────────────────────────
        trials = result.get("trial_reports", [])
        if trials:
            story.append(Paragraph(f"Per-Trial Accuracy Scores", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=PURPLE))
            story.append(Spacer(1, 6))

            header = [["Trial #", "Accuracy", "Text", "Color", "SSIM", "Icon", "Expiry", "Print", "Status"]]
            rows = []
            for t in trials:
                scores = t.get("scores", {})
                status = "✓ PASS" if t.get("passed") else "✗ FAIL"
                rows.append([
                    str(t.get("trial_idx", "?")),
                    f'{t.get("accuracy_score", 0):.1f}',
                    f'{scores.get("text", 0):.0f}',
                    f'{scores.get("color", 0):.0f}',
                    f'{scores.get("ssim", 0):.0f}',
                    f'{scores.get("icon_size", 0):.0f}',
                    f'{scores.get("expiry", 0):.0f}',
                    f'{scores.get("print_quality", 0):.0f}',
                    status,
                ])
            tbl = Table(header + rows, colWidths=[1.3*cm]*8 + [2*cm])
            style_cmds = [
                ("BACKGROUND", (0,0), (-1,0), NAVY),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("GRID", (0,0), (-1,-1), 0.3, SILVER),
                ("ALIGN", (1,1), (-1,-1), "CENTER"),
            ]
            for i, t in enumerate(trials, 1):
                bg = colors.HexColor("#FEE2E2") if not t.get("passed") else (colors.white if i % 2 else LIGHT)
                style_cmds.append(("BACKGROUND", (0,i), (-1,i), bg))
            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)
            story.append(Spacer(1, 14))

        # ── Issues found ──────────────────────────────────────────────────────
        all_critical = []
        all_warnings = []
        for t in trials:
            es = t.get("error_summary", {})
            for c in es.get("critical", [])[:5]:
                all_critical.append((t.get("trial_idx", "?"), c))
            for w in es.get("warning", [])[:5]:
                all_warnings.append((t.get("trial_idx", "?"), w))

        if all_critical:
            story.append(PageBreak())
            story.append(Paragraph(f"🔴 Critical Errors ({len(all_critical)})", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=RED))
            story.append(Spacer(1, 6))
            for trial_idx, err in all_critical[:25]:
                story.append(Paragraph(
                    f'<b>Trial #{trial_idx} • {err.get("category", "?")}:</b> {err.get("description", "")}',
                    ps("err", fontSize=9.5, textColor=RED)
                ))
                story.append(Spacer(1, 3))
            story.append(Spacer(1, 12))

        if all_warnings:
            story.append(Paragraph(f"🟡 Warnings ({len(all_warnings)})", h1))
            story.append(HRFlowable(width="100%", thickness=1, color=ORANGE))
            story.append(Spacer(1, 6))
            for trial_idx, w in all_warnings[:25]:
                story.append(Paragraph(
                    f'<b>Trial #{trial_idx} • {w.get("category", "?")}:</b> {w.get("description", "")}',
                    ps("warn", fontSize=9, textColor=colors.HexColor("#92400E"))
                ))
                story.append(Spacer(1, 3))

        # ── Annotated images ──────────────────────────────────────────────────
        for t in trials[:3]:  # Up to 3 trials with images
            ann_path = t.get("annotated_path")
            if ann_path and Path(ann_path).exists():
                story.append(PageBreak())
                story.append(Paragraph(
                    f"Trial #{t.get('trial_idx')} — Annotated Comparison", h1))
                story.append(HRFlowable(width="100%", thickness=1, color=PURPLE))
                story.append(Spacer(1, 8))
                story.append(RLImage(ann_path,
                                      width=A4[0] - 3*cm, height=(A4[0] - 3*cm) * 0.45))

        doc.build(story, onFirstPage=page_template, onLaterPages=page_template)
        log.info(f"Prepress PDF generated: {out_path}")
        return out_path

    except Exception as e:
        log.exception(f"Prepress PDF generation failed: {e}")
        return None


def generate_prepress_excel(job_id: str, config: dict, result: dict) -> Optional[Path]:
    """Generate prepress Excel report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_prepress_results.xlsx"

        wb = openpyxl.Workbook()
        header_fill = PatternFill("solid", fgColor="7C3AED")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        pass_fill = PatternFill("solid", fgColor="D1FAE5")
        fail_fill = PatternFill("solid", fgColor="FEE2E2")

        # Summary sheet
        ws = wb.active
        ws.title = "Decision Summary"
        decision = result.get("decision", "UNKNOWN")
        summary = [
            ["Job Reference", config.get("job_ref", job_id[:12])],
            ["Client", config.get("client_name", "")],
            ["Product", config.get("product_name", "")],
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["DECISION", decision],
            ["Reason", result.get("decision_reason", "")],
            ["Accuracy Score", f'{result.get("accuracy_score", 0):.1f}%'],
            ["Best Trial", f'{result.get("best_trial_score", 0):.1f}%'],
            ["Worst Trial", f'{result.get("worst_trial_score", 0):.1f}%'],
            ["Trials Inspected", result.get("trial_count", 0)],
            ["Min Required", f'{config.get("min_accuracy_for_go", 90)}%'],
            ["", ""],
            ["Waste Savings (USD)", f'${result.get("waste_savings", {}).get("estimated_savings_usd", 0):,.2f}'],
            ["Wasted m² avoided", result.get("waste_savings", {}).get("estimated_wasted_m2", 0)],
            ["Processing Time (ms)", result.get("processing_time_ms", 0)],
        ]
        for i, (k, v) in enumerate(summary, 1):
            ws.cell(i, 1, k).font = Font(bold=True)
            ws.cell(i, 2, str(v))
            if k == "DECISION":
                fill = pass_fill if v == "GO" else fail_fill
                ws.cell(i, 2).fill = fill
                ws.cell(i, 2).font = Font(bold=True, size=12)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # Per-trial sheet
        ws2 = wb.create_sheet("Per-Trial Scores")
        headers = ["Trial #", "Accuracy", "Pass/Fail", "Text", "Color", "SSIM",
                   "Icon Size", "Expiry", "Font Size", "Spell", "Print Quality",
                   "OCR Errors", "Defects", "Critical", "Warnings"]
        for col, h in enumerate(headers, 1):
            cell = ws2.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, t in enumerate(result.get("trial_reports", []), 2):
            scores = t.get("scores", {})
            es = t.get("error_summary", {})
            ws2.cell(i, 1, t.get("trial_idx"))
            ws2.cell(i, 2, t.get("accuracy_score"))
            ws2.cell(i, 3, "PASS" if t.get("passed") else "FAIL")
            ws2.cell(i, 4, scores.get("text"))
            ws2.cell(i, 5, scores.get("color"))
            ws2.cell(i, 6, scores.get("ssim"))
            ws2.cell(i, 7, scores.get("icon_size"))
            ws2.cell(i, 8, scores.get("expiry"))
            ws2.cell(i, 9, scores.get("font_size"))
            ws2.cell(i, 10, scores.get("spell"))
            ws2.cell(i, 11, scores.get("print_quality"))
            ws2.cell(i, 12, t.get("ocr_error_count", 0))
            ws2.cell(i, 13, t.get("defect_count", 0))
            ws2.cell(i, 14, es.get("critical_count", 0))
            ws2.cell(i, 15, es.get("warning_count", 0))
            fill = pass_fill if t.get("passed") else fail_fill
            for c in range(1, 16):
                ws2.cell(i, c).fill = fill
        for col in "ABCDEFGHIJKLMNO":
            ws2.column_dimensions[col].width = 12

        # Errors sheet
        ws3 = wb.create_sheet("All Issues")
        for col, h in enumerate(["Trial", "Severity", "Category", "Description"], 1):
            cell = ws3.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for t in result.get("trial_reports", []):
            es = t.get("error_summary", {})
            for sev_name, items in [("Critical", es.get("critical", [])),
                                      ("Warning", es.get("warning", [])),
                                      ("Info", es.get("info", []))]:
                for item in items:
                    ws3.cell(row, 1, t.get("trial_idx"))
                    ws3.cell(row, 2, sev_name)
                    ws3.cell(row, 3, item.get("category", ""))
                    ws3.cell(row, 4, item.get("description", ""))
                    if sev_name == "Critical":
                        for c in range(1, 5):
                            ws3.cell(row, c).fill = fail_fill
                    row += 1
        ws3.column_dimensions["A"].width = 8
        ws3.column_dimensions["B"].width = 12
        ws3.column_dimensions["C"].width = 15
        ws3.column_dimensions["D"].width = 60

        # Pantone sheet (if final design Pantones identified)
        final_pantones = result.get("final_pantones")
        if final_pantones and final_pantones.get("extracted_colors"):
            ws4 = wb.create_sheet("Final Design Pantones")
            for col, h in enumerate(["#", "PMS Code", "ΔE", "Confidence", "Area %",
                                       "RGB", "Hex", "Lab"], 1):
                cell = ws4.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill

            for i, c in enumerate(final_pantones["extracted_colors"], 2):
                ws4.cell(i, 1, i - 1)
                ws4.cell(i, 2, c.get("best_match_code"))
                ws4.cell(i, 3, c.get("best_match_delta_e"))
                ws4.cell(i, 4, c.get("match_confidence"))
                ws4.cell(i, 5, c.get("area_pct"))
                ws4.cell(i, 6, str(c.get("rgb")))
                ws4.cell(i, 7, c.get("hex"))
                ws4.cell(i, 8, str(c.get("lab")))
            for col in "ABCDEFGH":
                ws4.column_dimensions[col].width = 16

        wb.save(str(out_path))
        log.info(f"Prepress Excel generated: {out_path}")
        return out_path

    except Exception as e:
        log.exception(f"Prepress Excel generation failed: {e}")
        return None
