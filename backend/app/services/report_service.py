"""
Greenpack Pro — Report Generation Service
ReportLab PDF QC reports + openpyxl Excel export + Windows print
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.config import get_settings

log = logging.getLogger(__name__)
settings = get_settings()


def generate_pdf_report(
    job_id: str,
    config: dict,
    scores: dict,
    text_errors: list,
    color_result: dict,
    ssim_result: dict,
    barcode_result: list,
    annotated_path: Optional[Path] = None,
    ocr_timeout: bool = False,
) -> Optional[Path]:
    """Generate branded PDF QC report"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, Image as RLImage, PageBreak, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Output path
        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_report.pdf"

        # Colors
        NAVY = colors.HexColor("#0D1B2A")
        BLUE = colors.HexColor("#1A73E8")
        CYAN = colors.HexColor("#00C2CB")
        GREEN = colors.HexColor("#22A06B")
        RED = colors.HexColor("#E5383B")
        LIGHT = colors.HexColor("#F0F6FF")
        SILVER = colors.HexColor("#E8EEF4")

        # Override with client brand color if configured
        brand_color = colors.HexColor(config.get("brand_color", "#0D1B2A"))

        # Styles
        def ps(name, **kw):
            d = dict(fontName="Helvetica", fontSize=10, leading=14)
            d.update(kw)
            return ParagraphStyle(name, **d)

        title_style = ps("title", fontName="Helvetica-Bold", fontSize=24,
                         textColor=colors.white, alignment=TA_CENTER, leading=30)
        h1_style = ps("h1", fontName="Helvetica-Bold", fontSize=14, textColor=NAVY)
        h2_style = ps("h2", fontName="Helvetica-Bold", fontSize=11, textColor=NAVY)
        body_style = ps("body", fontSize=9.5, textColor=colors.HexColor("#333333"), leading=13)
        score_style = ps("score", fontName="Helvetica-Bold", fontSize=36,
                         alignment=TA_CENTER,
                         textColor=GREEN if scores["overall"] >= 75 else RED)

        # Page template with header/footer
        def page_template(canvas, doc):
            canvas.saveState()
            # Header bar
            canvas.setFillColor(brand_color)
            canvas.rect(0, A4[1] - 30, A4[0], 30, fill=1, stroke=0)
            canvas.setFont("Helvetica-Bold", 9)
            canvas.setFillColor(colors.white)
            canvas.drawString(1.5*cm, A4[1] - 20, "Greenpack Pro — Label QC Report")
            canvas.drawRightString(A4[0] - 1.5*cm, A4[1] - 20,
                f"Job: {config.get('job_ref', job_id[:8])}")
            # Footer
            canvas.setFillColor(SILVER)
            canvas.rect(0, 0, A4[0], 20, fill=1, stroke=0)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#666666"))
            canvas.drawString(1.5*cm, 6,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Greenpack Pro v1.0")
            canvas.drawRightString(A4[0] - 1.5*cm, 6, f"Page {doc.page}")
            canvas.restoreState()

        # Build document
        doc = SimpleDocTemplate(
            str(out_path), pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=1.5*cm
        )

        story = []

        # ── Cover Section ──────────────────────────────────────────────────────
        # Score header
        overall = scores["overall"]
        status = "✅ PASS" if overall >= 75 else "❌ FAIL"
        status_color = GREEN if overall >= 75 else RED

        cover_data = [[
            Paragraph(f'<font color="#FFFFFF"><b>{config.get("product_name", "Label Inspection")}</b></font>',
                      ps("cn", fontName="Helvetica-Bold", fontSize=16, textColor=colors.white, alignment=TA_CENTER)),
        ]]
        cover_tbl = Table(cover_data, colWidths=[A4[0] - 3*cm])
        cover_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), brand_color),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ]))
        story.append(cover_tbl)
        story.append(Spacer(1, 10))

        # Score + status
        score_data = [[
            Paragraph(f'{overall:.1f}', score_style),
            Paragraph(
                f'<b><font color="{"#22A06B" if overall >= 75 else "#E5383B"}">{status}</font></b>',
                ps("st", fontName="Helvetica-Bold", fontSize=20, alignment=TA_CENTER, leading=28)
            ),
        ]]
        score_tbl = Table(score_data, colWidths=[6*cm, 12*cm])
        score_tbl.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("BOX", (0, 0), (-1, -1), 1, SILVER),
        ]))
        story.append(score_tbl)
        story.append(Spacer(1, 12))

        # Sub-scores
        sub_data = [
            [Paragraph("OCR / Text", ps("sl", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Color", ps("sl", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Print Quality", ps("sl", fontName="Helvetica-Bold", fontSize=9)),
             Paragraph("Barcode", ps("sl", fontName="Helvetica-Bold", fontSize=9))],
            [Paragraph(f'{scores["ocr"]:.0f}', ps("sv", fontSize=18, fontName="Helvetica-Bold",
                       textColor=GREEN if scores["ocr"] >= 75 else RED, alignment=TA_CENTER)),
             Paragraph(f'{scores["color"]:.0f}', ps("sv2", fontSize=18, fontName="Helvetica-Bold",
                       textColor=GREEN if scores["color"] >= 75 else RED, alignment=TA_CENTER)),
             Paragraph(f'{scores["ssim"]:.0f}', ps("sv3", fontSize=18, fontName="Helvetica-Bold",
                       textColor=GREEN if scores["ssim"] >= 75 else RED, alignment=TA_CENTER)),
             Paragraph(f'{scores["barcode"]:.0f}', ps("sv4", fontSize=18, fontName="Helvetica-Bold",
                       textColor=GREEN if scores["barcode"] >= 75 else RED, alignment=TA_CENTER))],
        ]
        sub_tbl = Table(sub_data, colWidths=[(A4[0] - 3*cm) / 4] * 4)
        sub_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, SILVER),
        ]))
        story.append(sub_tbl)
        story.append(Spacer(1, 16))

        # Job details
        story.append(Paragraph("Inspection Details", h1_style))
        story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
        story.append(Spacer(1, 8))
        details_data = [
            ["Job Reference:", config.get("job_ref", job_id[:12]),
             "Client:", config.get("client_name", "—")],
            ["Product:", config.get("product_name", "—"),
             "Inspector:", config.get("inspector_name", "—")],
            ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M"),
             "Processing Time:", f'{config.get("processing_time_ms", 0) / 1000:.1f}s'],
        ]
        details_tbl = Table(details_data, colWidths=[3.5*cm, 7*cm, 2.5*cm, 4*cm])
        details_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, LIGHT]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(details_tbl)
        story.append(Spacer(1, 16))

        # ── OCR Errors ─────────────────────────────────────────────────────────
        if text_errors:
            story.append(Paragraph("OCR / Text Errors", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))

            err_header = [["Region", "Master Text", "Scanned Text", "Type", "Severity"]]
            err_rows = []
            for e in text_errors[:20]:  # Max 20 rows
                err_rows.append([
                    Paragraph(str(e.get("region_bbox", {}).get("x", "?"))[:15], body_style),
                    Paragraph(str(e.get("master_text", ""))[:30], body_style),
                    Paragraph(str(e.get("scan_text", ""))[:30], body_style),
                    Paragraph(e.get("type", ""), body_style),
                    Paragraph(e.get("severity", ""), body_style),
                ])

            err_tbl = Table(err_header + err_rows, colWidths=[2*cm, 5*cm, 5*cm, 2.5*cm, 2.5*cm])
            err_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                ("GRID", (0, 0), (-1, -1), 0.3, SILVER),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(err_tbl)
            story.append(Spacer(1, 12))

        # ── Barcode Results ─────────────────────────────────────────────────────
        if barcode_result:
            story.append(Paragraph("Barcode Verification", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))

            bc_header = [["Type", "Decoded Value", "Expected", "Match", "Check Digit", "Grade", "Status"]]
            bc_rows = []
            for b in barcode_result:
                status_txt = "✅ PASS" if b.get("pass") else "❌ FAIL"
                bc_rows.append([
                    b.get("type", ""),
                    str(b.get("decoded_value", "N/A"))[:20],
                    str(b.get("expected_value", "—"))[:20],
                    "Yes" if b.get("match") else "No",
                    "Yes" if b.get("check_digit_valid") else "No",
                    b.get("quality_grade", "?"),
                    status_txt,
                ])

            bc_tbl = Table(bc_header + bc_rows,
                          colWidths=[2*cm, 4.5*cm, 4.5*cm, 1.5*cm, 2*cm, 1.5*cm, 2*cm])
            bc_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                ("GRID", (0, 0), (-1, -1), 0.3, SILVER),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(bc_tbl)
            story.append(Spacer(1, 12))

        # ── Annotated Image ─────────────────────────────────────────────────────
        if annotated_path and Path(annotated_path).exists():
            story.append(PageBreak())
            story.append(Paragraph("Annotated Label Comparison", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=CYAN))
            story.append(Spacer(1, 8))
            img_width = A4[0] - 3*cm
            story.append(RLImage(str(annotated_path), width=img_width, height=img_width * 0.5))

        # Build PDF
        doc.build(story, onFirstPage=page_template, onLaterPages=page_template)
        log.info(f"PDF report generated: {out_path}")
        return out_path

    except Exception as e:
        log.error(f"PDF report generation failed: {e}")
        return None


def generate_excel_report(
    job_id: str,
    config: dict,
    scores: dict,
    text_errors: list,
    color_result: dict,
    ssim_result: dict,
    barcode_result: list,
) -> Optional[Path]:
    """Generate multi-sheet Excel QC report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_results.xlsx"

        wb = openpyxl.Workbook()

        # ── Summary Sheet ───────────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"

        header_fill = PatternFill("solid", fgColor="0D1B2A")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        pass_fill = PatternFill("solid", fgColor="D1FAE5")
        fail_fill = PatternFill("solid", fgColor="FEE2E2")

        summary_data = [
            ["Job Reference", config.get("job_ref", job_id[:12])],
            ["Client", config.get("client_name", "")],
            ["Product", config.get("product_name", "")],
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Overall Score", f'{scores["overall"]:.1f}'],
            ["Status", "PASS" if scores["overall"] >= 75 else "FAIL"],
            ["OCR Score", f'{scores["ocr"]:.1f}'],
            ["Color Score", f'{scores["color"]:.1f}'],
            ["SSIM Score", f'{scores["ssim"]:.1f}'],
            ["Barcode Score", f'{scores["barcode"]:.1f}'],
            ["OCR Errors", len(text_errors)],
            ["Barcode Failures", sum(1 for b in barcode_result if not b.get("pass"))],
            ["Defects Found", len(ssim_result.get("defects", []))],
        ]

        for row_idx, (key, val) in enumerate(summary_data, 1):
            ws_sum.cell(row_idx, 1, key).font = Font(bold=True)
            ws_sum.cell(row_idx, 2, str(val))
            if key == "Status":
                fill = pass_fill if val == "PASS" else fail_fill
                ws_sum.cell(row_idx, 2).fill = fill

        ws_sum.column_dimensions["A"].width = 20
        ws_sum.column_dimensions["B"].width = 30

        # ── OCR Errors Sheet ────────────────────────────────────────────────────
        ws_ocr = wb.create_sheet("OCR Errors")
        headers = ["Region X", "Region Y", "Master Text", "Scanned Text", "Type", "Severity"]
        for col, h in enumerate(headers, 1):
            cell = ws_ocr.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, err in enumerate(text_errors, 2):
            bbox = err.get("region_bbox", {})
            ws_ocr.cell(row_idx, 1, bbox.get("x", ""))
            ws_ocr.cell(row_idx, 2, bbox.get("y", ""))
            ws_ocr.cell(row_idx, 3, str(err.get("master_text", ""))[:100])
            ws_ocr.cell(row_idx, 4, str(err.get("scan_text", ""))[:100])
            ws_ocr.cell(row_idx, 5, err.get("type", ""))
            ws_ocr.cell(row_idx, 6, err.get("severity", ""))
            if err.get("severity") == "high":
                for col in range(1, 7):
                    ws_ocr.cell(row_idx, col).fill = fail_fill

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws_ocr.column_dimensions[col].width = 20

        # ── Barcode Results Sheet ───────────────────────────────────────────────
        ws_bc = wb.create_sheet("Barcode Results")
        bc_headers = ["Type", "Decoded Value", "Expected Value", "Match", "Check Digit", "Grade", "Status"]
        for col, h in enumerate(bc_headers, 1):
            cell = ws_bc.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, b in enumerate(barcode_result, 2):
            ws_bc.cell(row_idx, 1, b.get("type", ""))
            ws_bc.cell(row_idx, 2, str(b.get("decoded_value", "")))
            ws_bc.cell(row_idx, 3, str(b.get("expected_value", "")))
            ws_bc.cell(row_idx, 4, "Yes" if b.get("match") else "No")
            ws_bc.cell(row_idx, 5, "Yes" if b.get("check_digit_valid") else "No")
            ws_bc.cell(row_idx, 6, b.get("quality_grade", ""))
            status_val = "PASS" if b.get("pass") else "FAIL"
            ws_bc.cell(row_idx, 7, status_val)
            fill = pass_fill if b.get("pass") else fail_fill
            for col in range(1, 8):
                ws_bc.cell(row_idx, col).fill = fill

        # ── Color Results Sheet ─────────────────────────────────────────────────
        ws_col = wb.create_sheet("Color Results")
        col_headers = ["Zone", "Zone Name", "Mean ΔE", "Max ΔE", "% Over Threshold", "Threshold", "Status"]
        for col, h in enumerate(col_headers, 1):
            cell = ws_col.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, zone in enumerate(color_result.get("zone_results", []), 2):
            ws_col.cell(row_idx, 1, zone.get("zone_id", ""))
            ws_col.cell(row_idx, 2, zone.get("zone_name", ""))
            ws_col.cell(row_idx, 3, zone.get("mean_delta_e", ""))
            ws_col.cell(row_idx, 4, zone.get("max_delta_e", ""))
            ws_col.cell(row_idx, 5, zone.get("pct_out_of_tolerance", ""))
            ws_col.cell(row_idx, 6, zone.get("threshold", ""))
            status_val = "PASS" if zone.get("pass") else "FAIL"
            ws_col.cell(row_idx, 7, status_val)

        # ── Defects Sheet ───────────────────────────────────────────────────────
        ws_def = wb.create_sheet("Defects")
        def_headers = ["Type", "Severity", "X", "Y", "Width", "Height", "Area (px²)"]
        for col, h in enumerate(def_headers, 1):
            cell = ws_def.cell(1, col, h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, d in enumerate(ssim_result.get("defects", []), 2):
            bbox = d.get("bbox", {})
            ws_def.cell(row_idx, 1, d.get("type", ""))
            ws_def.cell(row_idx, 2, d.get("severity", ""))
            ws_def.cell(row_idx, 3, bbox.get("x", ""))
            ws_def.cell(row_idx, 4, bbox.get("y", ""))
            ws_def.cell(row_idx, 5, bbox.get("w", ""))
            ws_def.cell(row_idx, 6, bbox.get("h", ""))
            ws_def.cell(row_idx, 7, d.get("area_pixels", ""))

        wb.save(str(out_path))
        log.info(f"Excel report generated: {out_path}")
        return out_path

    except Exception as e:
        log.error(f"Excel report generation failed: {e}")
        return None


def print_report_windows(report_path: str, printer_name: str = None) -> bool:
    """Send PDF report directly to Windows printer"""
    try:
        import win32api
        import win32con
        win32api.ShellExecute(
            0, "print", report_path, None, ".", 0
        )
        log.info(f"Report sent to Windows printer: {report_path}")
        return True
    except ImportError:
        # Fallback: open with default application
        import os
        os.startfile(report_path, "print")
        return True
    except Exception as e:
        log.error(f"Print failed: {e}")
        return False
